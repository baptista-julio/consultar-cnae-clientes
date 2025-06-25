# Bibliotecas -------

import pandas as pd
import requests
import datetime
import oracledb
from unidecode import unidecode
from dotenv import load_dotenv
from os import environ
import os
import glob
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from time import sleep




# Configuração Obrigatória -------

caminho_biblioteca_oracle = r"C:\instantclient-basic-windows-64bits-23.7.0.25.01\instantclient_23_7"

autorizacao_receitaws = environ["AUTH_RECEITAWS"]

lista_cnaes_empresa = ["4673700", "4642702", "4649406","4651601", "4672900", "4679604","4679699", "4742300", "4753900"]

# Possíveis valores de PORTE: MEI, ME (Micro Empresa), EPP (Empresa de Pequeno Porte), DEMAIS (Médio e Grande Porte)
# IMPORTANTE: CNPJs são tratados como string para preservar zeros à esquerda e suportar futuros formatos com letras

cnpj_clientes = """
        SELECT 
            CODCLI, 
            REGEXP_REPLACE(cgcent, '[^0-9A-Za-z]', '') AS CNPJ -- OBRIGATÓRIO SER "CNPJ" MAIÚSCULO
        FROM 
            pcclient 
        WHERE 
            CODCLI IN (
                SELECT 
                    DISTINCT CODCLI 
                FROM 
                    pcpedc 
                WHERE 
                    DTFAT >= TRUNC(ADD_MONTHS(SYSDATE, -5), 'MM') 
                    AND DTFAT <= LAST_DAY(SYSDATE)
                    AND CODCLI NOT IN (
                        SELECT DISTINCT codcli 
                        FROM pcfilial 
                        WHERE codcli IS NOT NULL
                    )
            ) 
            AND TIPOFJ = 'J'
"""




# Configurações de Conexão com Banco de Dados Oracle -------

load_dotenv()

string_conexao_oracle = environ["USERNAME_ORACLE"] + '/'+ environ["PASSWORD_ORACLE"] + '@' + environ["HOST_ORACLE"] + ':' + environ["PORT_ORACLE"] + '/' + environ["SERVICE_NAME_ORACLE"]

oracledb.init_oracle_client(lib_dir = caminho_biblioteca_oracle)




# Formatos de data -------

data_ontem = (datetime.datetime.today() - datetime.timedelta(days=1)).strftime('%d-%m-%Y')

data_hoje = (datetime.datetime.today()).strftime('%d-%m-%Y')




# Funções -------

def consulta_oracle(consulta):
    connection = oracledb.connect(string_conexao_oracle) 
    cursor = connection.cursor()
    cursor.execute(
        f"""
            {consulta}
        """
    )
    resultado = cursor.fetchall()
    cabecalho = [desc[0] for desc in cursor.description]
    connection.close()
    dataframe = pd.DataFrame(data=resultado, columns=cabecalho)
    return dataframe


def consulta_receitaws(cnpj):
    days = 7 # Caso queira um dado mais atualizado, coloque 1 em vez de 7. Lembre-se da limitação do seu plano no site ReceitaWS.
    url = f"https://receitaws.com.br/v1/cnpj/{cnpj}/days/{days}"
    headers = {
        'Accept': 'application/json',
        'Authorization': autorizacao_receitaws
    }
    try:
        response = requests.get(url, headers=headers, timeout=10)
        if response.status_code == 200:
            return response.json()
        else:
            return f"Erro ao consultar CNPJ: {response.status_code} - {response.text}"
    except requests.exceptions.RequestException as e:
        return f"Erro na requisição: {e}"


def buscar_arquivo_excel_mais_recente():
    """Busca o arquivo Excel mais recente com padrão 'CNPJ_consulta_incremental_*.xlsx'"""
    arquivos = glob.glob('CNPJ_consulta_incremental_*.xlsx')
    if arquivos:
        # Ordena os arquivos por data de modificação (mais recente primeiro)
        arquivos_ordenados = sorted(arquivos, key=os.path.getmtime, reverse=True)
        return arquivos_ordenados[0]
    return None


def anexar_ao_excel(arquivo_excel, dados_consultados, dados_erros, dados_restantes):
    """Anexa dados ao arquivo Excel existente ou cria um novo"""
    
    # Se o arquivo não existe, criar com pandas normalmente
    if not os.path.exists(arquivo_excel):
        with pd.ExcelWriter(arquivo_excel, engine='xlsxwriter') as writer:
            workbook = writer.book
            text_format = workbook.add_format({'num_format': '@'})  # Formato de texto
            
            # Aba Consultados
            if not dados_consultados.empty:
                dados_consultados['CNPJ'] = dados_consultados['CNPJ'].astype(str)
                dados_consultados.to_excel(writer, sheet_name='Consultados', index=False)
                worksheet = writer.sheets['Consultados']
                col_idx = dados_consultados.columns.get_loc('CNPJ')
                worksheet.set_column(col_idx, col_idx, 15, text_format)
            else:
                pd.DataFrame(columns=['CODCLI', 'CNPJ', 'NOME EMPRESA', 'NOME FANTASIA', 'PORTE', 
                                    'SITUACAO CNPJ', 'CNAE', 'DESCRICAO CNAE', 'TIPO CNAE', 
                                    'IGUALDADE', 'QTD IGUAL', 'QTD DIFERENTE', 'COMANDO INSERT']).to_excel(
                    writer, sheet_name='Consultados', index=False)
            
            # Aba Restantes
            if not dados_restantes.empty:
                dados_restantes['CNPJ'] = dados_restantes['CNPJ'].astype(str)
            dados_restantes.to_excel(writer, sheet_name='Restantes', index=False)
            if not dados_restantes.empty:
                worksheet = writer.sheets['Restantes']
                col_idx = dados_restantes.columns.get_loc('CNPJ')
                worksheet.set_column(col_idx, col_idx, 15, text_format)
            
            # Aba Erros Consulta
            if not dados_erros.empty:
                dados_erros['CNPJ'] = dados_erros['CNPJ'].astype(str)
                dados_erros.to_excel(writer, sheet_name='Erros Consulta', index=False)
                worksheet = writer.sheets['Erros Consulta']
                col_idx = dados_erros.columns.get_loc('CNPJ')
                worksheet.set_column(col_idx, col_idx, 15, text_format)
            else:
                pd.DataFrame(columns=['CODCLI', 'CNPJ', 'ERRO']).to_excel(
                    writer, sheet_name='Erros Consulta', index=False)
        return
    
    # Se o arquivo existe, usar openpyxl para anexar
    wb = load_workbook(arquivo_excel)
    
    # Anexar dados consultados
    if not dados_consultados.empty and 'Consultados' in wb.sheetnames:
        ws = wb['Consultados']
        # Encontrar a primeira linha vazia
        primeira_linha_vazia = ws.max_row + 1
        
        # Adicionar os novos dados
        for r_idx, row in enumerate(dataframe_to_rows(dados_consultados, index=False, header=False), primeira_linha_vazia):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Anexar erros
    if not dados_erros.empty and 'Erros Consulta' in wb.sheetnames:
        ws = wb['Erros Consulta']
        primeira_linha_vazia = ws.max_row + 1
        
        for r_idx, row in enumerate(dataframe_to_rows(dados_erros, index=False, header=False), primeira_linha_vazia):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Substituir completamente a aba Restantes (não anexar)
    if 'Restantes' in wb.sheetnames:
        ws = wb['Restantes']
        # Limpar conteúdo mantendo cabeçalho
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None
        
        # Adicionar novos restantes
        if not dados_restantes.empty:
            for r_idx, row in enumerate(dataframe_to_rows(dados_restantes, index=False, header=False), 2):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Salvar
    wb.save(arquivo_excel)
    wb.close()


def carregar_restantes_do_excel(arquivo_excel):
    """Carrega a aba 'Restantes' do arquivo Excel se existir"""
    try:
        # Força a leitura da coluna CNPJ como string para preservar zeros à esquerda
        df_restantes = pd.read_excel(arquivo_excel, sheet_name='Restantes', dtype={'CNPJ': str})
        if not df_restantes.empty and 'CODCLI' in df_restantes.columns and 'CNPJ' in df_restantes.columns:
            print(f"Carregando {len(df_restantes)} CNPJs restantes do arquivo: {arquivo_excel}")
            return df_restantes
    except Exception as e:
        print(f"Erro ao carregar aba 'Restantes': {e}")
    return None


def contar_registros_processados(arquivo_excel):
    """Conta quantos registros já foram processados no arquivo"""
    contagem = 0
    try:
        # Contar consultados
        df_consultados = pd.read_excel(arquivo_excel, sheet_name='Consultados', dtype={'CNPJ': str})
        cnpjs_consultados = set(df_consultados['CNPJ'].unique()) if not df_consultados.empty else set()
        
        # Contar erros
        df_erros = pd.read_excel(arquivo_excel, sheet_name='Erros Consulta', dtype={'CNPJ': str})
        cnpjs_erros = set(df_erros['CNPJ'].unique()) if not df_erros.empty else set()
        
        # União dos CNPJs processados (consultados + erros)
        contagem = len(cnpjs_consultados.union(cnpjs_erros))
    except:
        pass
    
    return contagem




# MAIN -------

# Nome do arquivo principal
arquivo_principal = f'CNPJ_consulta_incremental_{data_hoje}.xlsx'

# Verificar se existe arquivo do dia ou outro arquivo recente
arquivo_excel_existente = None
if os.path.exists(arquivo_principal):
    arquivo_excel_existente = arquivo_principal
else:
    arquivo_excel_existente = buscar_arquivo_excel_mais_recente()

df_clientes_da_empresa = None
contagem_inicial = 0

if arquivo_excel_existente:
    print(f"\nArquivo encontrado: {arquivo_excel_existente}")
    
    # Carregar restantes se existirem
    df_restantes_anterior = carregar_restantes_do_excel(arquivo_excel_existente)
    if df_restantes_anterior is not None and not df_restantes_anterior.empty:
        df_clientes_da_empresa = df_restantes_anterior
        contagem_inicial = contar_registros_processados(arquivo_excel_existente)
        print(f"Já foram processados {contagem_inicial} CNPJs")
        print(f"Continuando processamento a partir dos {len(df_clientes_da_empresa)} CNPJs restantes\n")

# Se não encontrou restantes ou arquivo anterior, consulta o banco
if df_clientes_da_empresa is None:
    print("Consultando CNPJs do banco de dados...")
    df_clientes_da_empresa = consulta_oracle(cnpj_clientes)

# Garantir que CNPJ seja tratado como string
df_clientes_da_empresa['CNPJ'] = df_clientes_da_empresa['CNPJ'].astype(str).str.strip()

# DataFrames para acumular dados da sessão atual
df_consultados_sessao = pd.DataFrame()
df_erros_sessao = pd.DataFrame()
restantes = df_clientes_da_empresa.copy()

contagem = contagem_inicial

for idx, row in df_clientes_da_empresa.iterrows():

    cnpj_atual = row['CNPJ']
    dados = consulta_receitaws(cnpj_atual)

    contagem += 1
    print(f"{contagem}. CNPJ: {row['CNPJ']}")

    if isinstance(dados, dict):
        # Verifica se houve erro na resposta da API (mesmo sendo dict)
        if 'status' in dados and dados['status'] == 'ERROR':
            erro_msg = dados.get('message', 'Erro desconhecido na API')
            df_erro_temp = pd.DataFrame([{
                'CODCLI': row['CODCLI'],
                'CNPJ': cnpj_atual,
                'ERRO': erro_msg
            }])
            df_erros_sessao = pd.concat([df_erros_sessao, df_erro_temp], ignore_index=True)
            print(f"Erro na consulta do CNPJ {cnpj_atual}: {erro_msg}")
        else:
            # Processa normalmente se não houver erro
            cnaes = [dados['atividade_principal'][0]] + dados.get('atividades_secundarias', [])
            registros_cnpj = []

            for atividade in cnaes:
                cnae_code = atividade['code'].replace(".", "").replace("-", "")
                igualdade = 'IGUAL' if cnae_code in lista_cnaes_empresa else 'DIFERENTE'

                registros_cnpj.append({
                    'CODCLI': row['CODCLI'],
                    'CNPJ': cnpj_atual,
                    'NOME EMPRESA': unidecode(dados.get('nome', 'Nome não encontrado')).upper(),
                    'NOME FANTASIA': unidecode(dados.get('fantasia', 'Nome fantasia não encontrado')).upper(),
                    'PORTE': dados.get('porte', 'PORTE NAO INFORMADO').upper(),
                    'SITUACAO CNPJ': dados.get('situacao', 'Situação não encontrada').upper(),
                    'CNAE': cnae_code,
                    'DESCRICAO CNAE': unidecode(atividade.get('text', 'Descrição não encontrada')).upper(),
                    'TIPO CNAE': 'PRIMARIO' if atividade == dados['atividade_principal'][0] else 'SECUNDARIO',
                    'IGUALDADE': igualdade
                })

            # Converter para DataFrame imediatamente
            df_temp = pd.DataFrame(registros_cnpj)
            df_consultados_sessao = pd.concat([df_consultados_sessao, df_temp], ignore_index=True)

    else:
        # Adicionar ao DataFrame de erros
        erro_msg = str(dados) if dados else "Retorno da consulta não é do tipo dict"
        df_erro_temp = pd.DataFrame([{
            'CODCLI': row['CODCLI'],
            'CNPJ': cnpj_atual,
            'ERRO': erro_msg
        }])
        df_erros_sessao = pd.concat([df_erros_sessao, df_erro_temp], ignore_index=True)
        
        print("\n---\n")
        print(f"Erro ao consultar o CNPJ: {cnpj_atual}")
        print(f"\nResposta recebida:\n{dados}")
        print("\n---\n")

    # Atualizar restantes
    restantes = restantes[restantes['CNPJ'] != cnpj_atual]

    # Salvar incrementalmente a cada 10 consultas ou no fim
    if (idx + 1) % 10 == 0 or (idx + 1) == len(df_clientes_da_empresa):
        df_consultados_para_salvar = df_consultados_sessao.copy()
        
        if not df_consultados_para_salvar.empty:
            # Calcular QTD IGUAL e DIFERENTE
            df_consultados_para_salvar['QTD IGUAL'] = df_consultados_para_salvar.groupby('CNPJ')['IGUALDADE'].transform(lambda x: (x == 'IGUAL').sum())
            df_consultados_para_salvar['QTD DIFERENTE'] = df_consultados_para_salvar.groupby('CNPJ')['IGUALDADE'].transform(lambda x: (x == 'DIFERENTE').sum())

            # Comando SQL completo
            df_consultados_para_salvar['COMANDO INSERT'] = df_consultados_para_salvar.apply(
                lambda row: f"INSERT INTO tabela (CODCLI, CNPJ, NOME_EMPRESA, NOME_FANTASIA, PORTE, SITUACAO_CNPJ, CNAE, DESCRICAO_CNAE, TIPO_CNAE, IGUALDADE, QTD_IGUAL, QTD_DIFERENTE) VALUES ('{row['CODCLI']}', '{row['CNPJ']}', '{row['NOME EMPRESA']}', '{row['NOME FANTASIA']}', '{row['PORTE']}', '{row['SITUACAO CNPJ']}', '{row['CNAE']}', '{row['DESCRICAO CNAE']}', '{row['TIPO CNAE']}', '{row['IGUALDADE']}', '{row['QTD IGUAL']}', '{row['QTD DIFERENTE']}');",
                axis=1
            )

        # Anexar ao arquivo Excel
        anexar_ao_excel(arquivo_principal, df_consultados_para_salvar, df_erros_sessao, restantes)
        
        print(f'Arquivo atualizado: {arquivo_principal}')
        
        # Exibir resumo
        print(f'\nResumo da sessão atual:')
        print(f'- Consultados nesta sessão: {df_consultados_sessao["CNPJ"].nunique() if not df_consultados_sessao.empty else 0}')
        print(f'- Erros nesta sessão: {len(df_erros_sessao)}')
        print(f'- Total restante: {len(restantes)}\n')

        sleep(3)
        
        # Limpar DataFrames da sessão após salvar
        df_consultados_sessao = pd.DataFrame()
        df_erros_sessao = pd.DataFrame()